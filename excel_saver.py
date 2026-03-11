# excel_saver.py
# WORKER 6 — Saves everything to Excel files with timestamps

import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

from config.settings import EXCEL_DIR
from config.logger import setup_logger

logger = setup_logger("excel")


class ExcelSaver:
    """
    Saves all data to organized Excel files.
    
    Creates TWO files:
    1. Daily Report: One file per day with all details
    2. Master Database: Cumulative file that grows every day
    """
    
    def __init__(self):
        self.today = datetime.now().strftime("%Y-%m-%d")
        self.timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        self.daily_file = os.path.join(
            EXCEL_DIR, f"daily_report_{self.today}.xlsx"
        )
        self.master_file = os.path.join(
            EXCEL_DIR, "master_database.xlsx"
        )
        
        # Pretty styles for Excel
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.header_fill = PatternFill(
            start_color="2F5496", end_color="2F5496", fill_type="solid"
        )
        self.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
    
    def _make_pretty(self, ws, num_cols):
        """Apply nice formatting to header row"""
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = self.border
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    def save_scraped_news(self, articles):
        """Save raw scraped news to Excel"""
        logger.info(f"💾 Saving {len(articles)} scraped articles to Excel...")
        
        data = []
        for i, a in enumerate(articles, 1):
            data.append({
                'S.No': i,
                'Date': a.get('date', self.today),
                'Scraped At': a.get('scraped_at', self.timestamp),
                'Source': a.get('source', ''),
                'Title': a.get('title', ''),
                'Link': a.get('link', ''),
                'Content Length': len(a.get('content', '')),
            })
        
        df = pd.DataFrame(data)
        
        with pd.ExcelWriter(self.daily_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Raw News', index=False)
        
        wb = load_workbook(self.daily_file)
        self._make_pretty(wb['Raw News'], len(df.columns))
        wb.save(self.daily_file)
        
        logger.info(f"   Saved to: {self.daily_file}")
    
    def save_filtered_news(self, filtered_articles, stats):
        """Save filtered news with AI analysis to Excel"""
        logger.info(f"💾 Saving {len(filtered_articles)} filtered articles...")
        
        data = []
        for i, a in enumerate(filtered_articles, 1):
            ev = a.get('evaluation', {})
            data.append({
                'S.No': i,
                'Date': a.get('date', self.today),
                'Filtered At': a.get('filtered_at', self.timestamp),
                'Category': ev.get('category', ''),
                'Importance': ev.get('importance', 0),
                'Title': a.get('title', ''),
                'Source': a.get('source', ''),
                'Summary': ev.get('one_line_summary', ''),
                'Key Facts': ', '.join(ev.get('key_facts', [])),
                'Link': a.get('link', ''),
            })
        
        df = pd.DataFrame(data)
        
        wb = load_workbook(self.daily_file)
        
        # Write Filtered News sheet
        ws = wb.create_sheet('Filtered News')
        headers = list(df.columns)
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        for row_idx, row in enumerate(data, 2):
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row.get(h, ''))
                cell.border = self.border
                cell.alignment = Alignment(wrap_text=True)
        self._make_pretty(ws, len(headers))
        
        # Write Statistics sheet
        ws2 = wb.create_sheet('Statistics')
        stat_rows = [
            ['Metric', 'Value', 'Timestamp'],
            ['Date', self.today, self.timestamp],
            ['Total Scraped', stats['total'], ''],
            ['Keyword Passed', stats['keyword_passed'], ''],
            ['Keyword Skipped', stats['keyword_skipped'], ''],
            ['AI Relevant', stats['ai_relevant'], ''],
            ['AI Not Relevant', stats['ai_not_relevant'], ''],
            ['AI Errors', stats['ai_errors'], ''],
            ['Final Selected', stats['final_selected'], ''],
        ]
        for r, row in enumerate(stat_rows, 1):
            for c, val in enumerate(row, 1):
                cell = ws2.cell(row=r, column=c, value=val)
                cell.border = self.border
        self._make_pretty(ws2, 3)
        
        wb.save(self.daily_file)
        logger.info(f"   Saved filtered news + statistics")
    
    def save_posts(self, post_data):
        """Save generated post content to Excel"""
        if not post_data:
            return
        
        logger.info(f"💾 Saving {len(post_data)} posts to Excel...")
        
        wb = load_workbook(self.daily_file)
        ws = wb.create_sheet('Generated Posts')
        
        headers = list(post_data[0].keys())
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        
        for row_idx, row in enumerate(post_data, 2):
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(
                    row=row_idx, column=col_idx,
                    value=str(row.get(h, ''))
                )
                cell.border = self.border
                cell.alignment = Alignment(wrap_text=True)
        
        self._make_pretty(ws, len(headers))
        wb.save(self.daily_file)
    
    def save_quiz(self, questions):
        """Save quiz questions to Excel"""
        logger.info(f"💾 Saving {len(questions)} quiz questions to Excel...")
        
        data = []
        for q in questions:
            data.append({
                'Q.No': q.get('question_number', ''),
                'Date': q.get('date', self.today),
                'Time': q.get('generated_at', self.timestamp),
                'Category': q.get('category', ''),
                'Question': q.get('question', ''),
                'Option A': q.get('option_a', ''),
                'Option B': q.get('option_b', ''),
                'Option C': q.get('option_c', ''),
                'Option D': q.get('option_d', ''),
                'Answer': q.get('correct_answer', ''),
                'Explanation': q.get('explanation', ''),
                'Source Article': q.get('source_article', ''),
            })
        
        wb = load_workbook(self.daily_file)
        ws = wb.create_sheet('Quiz Questions')
        
        headers = list(data[0].keys())
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        
        for row_idx, row in enumerate(data, 2):
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(
                    row=row_idx, column=col_idx,
                    value=str(row.get(h, ''))
                )
                cell.border = self.border
                cell.alignment = Alignment(wrap_text=True)
        
        self._make_pretty(ws, len(headers))
        wb.save(self.daily_file)
    
    def save_posting_log(self, post_type, status, details=""):
        """Record when something was posted to Telegram"""
        wb = load_workbook(self.daily_file)
        
        if 'Posting Log' not in wb.sheetnames:
            ws = wb.create_sheet('Posting Log')
            for col, h in enumerate(
                ['Date', 'Time', 'Type', 'Status', 'Details'], 1
            ):
                ws.cell(row=1, column=col, value=h)
            self._make_pretty(ws, 5)
        else:
            ws = wb['Posting Log']
        
        row = ws.max_row + 1
        values = [
            self.today,
            datetime.now().strftime("%H:%M:%S"),
            post_type,
            status,
            details
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = self.border
        
        wb.save(self.daily_file)
    
    def update_master(self, filtered_articles, questions):
        """Add today's data to the master cumulative database"""
        logger.info("💾 Updating master database...")
        
        # Prepare news data
        news_rows = []
        for a in filtered_articles:
            ev = a.get('evaluation', {})
            news_rows.append({
                'Date': a.get('date', self.today),
                'Time': a.get('filtered_at', self.timestamp),
                'Category': ev.get('category', ''),
                'Importance': ev.get('importance', 0),
                'Title': a.get('title', ''),
                'Source': a.get('source', ''),
                'Summary': ev.get('one_line_summary', ''),
                'Key Facts': ', '.join(ev.get('key_facts', [])),
                'Link': a.get('link', ''),
            })
        
        # Prepare quiz data
        quiz_rows = []
        for q in questions:
            quiz_rows.append({
                'Date': q.get('date', self.today),
                'Time': q.get('generated_at', ''),
                'Category': q.get('category', ''),
                'Question': q.get('question', ''),
                'A': q.get('option_a', ''),
                'B': q.get('option_b', ''),
                'C': q.get('option_c', ''),
                'D': q.get('option_d', ''),
                'Answer': q.get('correct_answer', ''),
                'Explanation': q.get('explanation', ''),
            })
        
        news_df = pd.DataFrame(news_rows)
        quiz_df = pd.DataFrame(quiz_rows)
        
        # If master file exists, append to it
        if os.path.exists(self.master_file):
            try:
                old_news = pd.read_excel(self.master_file, sheet_name='All News')
                news_df = pd.concat([old_news, news_df], ignore_index=True)
            except:
                pass
            try:
                old_quiz = pd.read_excel(self.master_file, sheet_name='All Quiz')
                quiz_df = pd.concat([old_quiz, quiz_df], ignore_index=True)
            except:
                pass
        
        # Write master file
        with pd.ExcelWriter(self.master_file, engine='openpyxl') as writer:
            news_df.to_excel(writer, sheet_name='All News', index=False)
            quiz_df.to_excel(writer, sheet_name='All Quiz', index=False)
        
        # Make it pretty
        wb = load_workbook(self.master_file)
        for sheet in wb.sheetnames:
            self._make_pretty(wb[sheet], wb[sheet].max_column)
        wb.save(self.master_file)
        
        logger.info(
            f"   Master DB: {len(news_df)} total news, "
            f"{len(quiz_df)} total questions"
        )